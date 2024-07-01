import json
import os
from tkinter import messagebox
from reg_qpcr import InvalidExcelLayoutException
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import xlrd

'''
Rules:
1. One strip can only have one target (absent of improvement 1)
2. Naming conventions must be consistent and correct, e.g. cannot label NTC as 10 input
3. Can only pick one set of reference samples
'''

'''
Improvements:
1. Make an interface to allow users to choose what they want compared if they have weird requests (advanced option)
2. Add option for using different target names
'''

def default_settings():
    """Return the default settings."""
    return {
        "REF_MODE": "Exclude Other References",
        "DECIMAL_PLACES": 100,
        "BATCH_TEXT_COLOR": "ff003822",
        "BATCH_FILL_COLOR": "ffedffdc",
        "AVERAGE_BATCHES_NAME": "Average (Do not delete this if you want continuous input)",
        "PERCENTAGE_INPUT": 1,
        "SAMPLE_COLUMN": ["sample", "sample name", "name"],
        "WELL_COLUMN": ["well"],
        "GENE_COLUMN": ["gene", "target", "target name", "detector", "primer/probe", "primer", "probe"],
        "CT_COLUMN": ["ct", "cq", "cт"]
    }


def get_settings():
    """Load settings from JSON file. If the file doesn't exist, return default settings."""
    try:
        with open("settings.json", "r") as file:
            return json.load(file)
    except FileNotFoundError:
        return default_settings()


settings = get_settings()

PERCENTAGE_INPUT = settings["PERCENTAGE_INPUT"]

SAMPLE_COLUMN = settings["SAMPLE_COLUMN"]
WELL_COLUMN = settings["WELL_COLUMN"]
GENE_COLUMN = settings["GENE_COLUMN"]
CT_COLUMN = settings["CT_COLUMN"]


def get_data(wb_file, what_sort):
    label_row = -1
    raw_data = []
    samples = []

    file_extension = os.path.splitext(wb_file)[1]

    if file_extension == '.xls':
        wb = xlrd.open_workbook(wb_file)
        read_ws = wb.sheet_by_index(0)
        row_count, col_count = read_ws.nrows, read_ws.ncols

        for row_num in range(0, row_count):
            cell = read_ws.cell(row_num, 0)
            if cell.value.lower() in WELL_COLUMN:
                label_row = row_num
                for num1 in range(0, col_count):
                    cell = read_ws.cell(label_row, num1)
                    if cell.value.lower() in WELL_COLUMN:
                        for i in range(label_row + 1, row_count):
                            well_number = read_ws.cell(i, num1).value
                            raw_data.append([well_number])

                    if cell.value.lower() in SAMPLE_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count):
                            cell = read_ws.cell(i, num1)
                            sample_name = cell.value
                            raw_data[counter].append(sample_name)
                            counter += 1

                    if cell.value.lower() in GENE_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count):
                            cell = read_ws.cell(i, num1)
                            target_name = cell.value
                            raw_data[counter].append(target_name)
                            counter += 1

                    if cell.value.lower() in CT_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count):
                            cell = read_ws.cell(i, num1)
                            ct = cell.value
                            raw_data[counter].append(ct)
                            counter += 1

        if label_row == -1:
            raise InvalidExcelLayoutException("Invalid Excel File Layout for file: {}".format(wb_file))

    elif file_extension == ".xlsx":
        wb = openpyxl.load_workbook(wb_file)
        read_ws = wb.active
        row_count = read_ws.max_row
        col_count = read_ws.max_column

        for row_num in range(1, row_count + 1):   # openpyxl is 1-indexed
            cell = read_ws.cell(row=row_num, column=1)
            if cell.value and cell.value.lower() in WELL_COLUMN:
                label_row = row_num
                for num1 in range(1, col_count+1):
                    cell = read_ws.cell(row=label_row, column=num1)
                    if cell.value and cell.value.lower() in WELL_COLUMN:
                        for i in range(label_row + 1, row_count+1):
                            cell = read_ws.cell(row=i, column=num1)
                            well_number = cell.value
                            raw_data.append([well_number])

                    if cell.value and cell.value.lower() in SAMPLE_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count+1):
                            cell = read_ws.cell(row=i, column=num1)
                            sample_name = cell.value
                            raw_data[counter].append(sample_name)
                            counter += 1

                    if cell.value and cell.value.lower() in GENE_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count+1):
                            cell = read_ws.cell(row=i, column=num1)
                            target_name = cell.value
                            raw_data[counter].append(target_name)
                            counter += 1

                    if cell.value and cell.value.lower() in CT_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count+1):
                            cell = read_ws.cell(row=i, column=num1)
                            ct = cell.value
                            raw_data[counter].append(ct)
                            counter += 1
                            print(f"{i}, {num1}")

        if label_row == -1:
            raise InvalidExcelLayoutException("Invalid Excel File Layout for file: {}".format(wb_file))

    # Clean the data
    for point in raw_data.copy():
        if point[1] == "":
            point[1] = "Unknown Sample"

        if point[3] == "":
            raw_data.remove(point)

    for datum in raw_data:
        if datum[2] == "":
            messagebox.showerror("Error", "One or more of the wells do not have a target")
            raise Exception("One or more of the wells do not have a target")

    # The consequence of this method is that one strip can only have one target
    data = []
    if what_sort == "Horizontal":
        # sort datapoints according to which row they are in (horizontal sort)
        prev_letter = ""
        current_num = -1
        for num in range(len(raw_data)):
            if raw_data[num][0][0] != prev_letter:
                current_num += 1
                data.append([raw_data[num]])
                prev_letter = raw_data[num][0][0]
            else:
                data[current_num].append(raw_data[num])
                prev_letter = raw_data[num][0][0]

    elif what_sort == "Vertical":
        # sort datapoints according to which column they are in (vertical sort)
        prev_number = ""
        data = []
        current_num = -1
        for num in range(len(raw_data)):
            if raw_data[num][0][1] != prev_number:
                current_num += 1
                data.append([raw_data[num]])
                prev_number = raw_data[num][0][1]
            else:
                data[current_num].append(raw_data[num])
                prev_number = raw_data[num][0][1]

    for datum in data:
        for name in datum:
            sample_name = name[1]
            if sample_name not in samples:
                samples.append(sample_name)

    return data, samples


def write_wb(data, reference_targets: list, graph_targets: list, output_filename):
    def make_cell(col_num, row_num, input_val, bold=False):
        if isinstance(input_val, int):
            ws[f"{alphabet[col_num]}{row_num}"] = f"={input_val}"
        else:
            ws[f"{alphabet[col_num]}{row_num}"] = f"{input_val}"

        if bold:
            bold_cell = ws[f"{alphabet[col_num]}{row_num}"]
            bold_cell.font = Font(bold=True)

    labels = ["Sample", "Raw", "Average", "delta Ct", "% of Input"]
    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']

    for datum in data:
        for point in datum:
            if len(point) > 0:
                # Make undetected = 0
                if not isinstance(point[3], float) and not isinstance(point[3], int):
                    point[3] = ""

    # Dictionary to store the assimilated lists
    result_dict = {}

    # Loop through each list
    for sublist in graph_targets:
        key = sublist[0]
        if key not in result_dict:
            result_dict[key] = sublist
        else:
            result_dict[key].extend(sublist[1:])

    graph_targets_true = list(result_dict.values())

    wb = Workbook()
    # Remove the default "Sheet"
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    for target in graph_targets_true:

        ws = wb.create_sheet(f"{target[0]}")

        cell_letter, cell_num = 0, 0
        sample_col = 0
        sample_row = 0
        num_labels = 5
        spacing = 1
        sample_name = None
        target_name = ""
        adj_input_coor = []
        percent_inputs = []
        left_current_row = 1
        right_current_row = 1

        # log(100,2)
        make_cell(-2, 3, f"=LOG({100 / PERCENTAGE_INPUT},2)")
        make_cell(-2, 2, f"{PERCENTAGE_INPUT}%={100 / PERCENTAGE_INPUT} folds")

        for outside in range(len(data)):
            # works for the IRS or something idk
            internal_counter = 0
            # works for cell placement
            counter = 0
            if outside % 2 == 0:
                for inside in range(len(data[outside])):
                    if internal_counter == 0:
                        # Target Name
                        target_name = data[outside][inside][2]
                        ws[f"{alphabet[0]}{left_current_row}"] = target_name
                        ws[f"{alphabet[0]}{left_current_row}"].font = Font(bold=True)

                        for i in range(num_labels):
                            cell_letter = i + 1
                            cell_num = 1 + left_current_row
                            ws[f"{alphabet[cell_letter]}{cell_num}"] = labels[i]
                            ws[f"{alphabet[cell_letter]}{cell_num}"].font = Font(bold=True)

                        left_current_row += len(data[outside]) + 4
                        # +1 for spacing
                        last_hori_cell = [cell_letter + 1, cell_num]

                        sample_col = last_hori_cell[0] - num_labels
                        sample_row = last_hori_cell[1] + 1

                    # establish adjusted input cell location so that you can reference there while making the rest
                    if internal_counter == 0:
                        make_cell(sample_col, sample_row + len(data[outside]), "Adjusted Input (100%)")
                        adj_input_coor = [sample_col + 1, sample_row + len(data[outside])]

                    # prev_sample = data[outside][inside-1][1]
                    current_sample = data[outside][inside][1]

                    if inside + 1 < len(data[outside]):
                        next_sample = data[outside][inside + 1][1]
                    else:
                        next_sample = "!@#&I#!@^&*!@#*&^!@#()*&"

                    if next_sample != current_sample:
                        # Sample Name
                        sample_name = data[outside][inside][1]
                        ws[f"{alphabet[sample_col]}{sample_row - counter}"] = sample_name
                        # Average
                        ws[
                            f"{alphabet[sample_col + 2]}{sample_row - counter}"] = f"=AVERAGE({alphabet[sample_col + 1]}{sample_row - counter}: {alphabet[sample_col + 1]}{sample_row})"
                        # deltact
                        ws[
                            f"{alphabet[sample_col + 3]}{sample_row - counter}"] = f"={alphabet[adj_input_coor[0]]}{adj_input_coor[1]}- {alphabet[sample_col + 2]}{sample_row - counter}"

                        # % Input
                        make_cell(sample_col + 4, sample_row - counter,
                                  f"=100*2^{alphabet[sample_col + 3]}{sample_row - counter}")
                        percent_inputs.append([sample_name, target_name, [sample_col + 4, sample_row - counter]])

                        # now add values to adj input
                        for ref in reference_targets:
                            if ref == data[outside][inside][1]:
                                ave_input = [sample_col + 2, sample_row - counter]
                                make_cell(adj_input_coor[0], adj_input_coor[1],
                                          f"={alphabet[ave_input[0]]}{ave_input[1]}-{alphabet[-2]}3")
                        counter = 0
                    else:
                        counter += 1

                    internal_counter += 1

                    # Raw
                    ws[f"{alphabet[sample_col + 1]}{sample_row}"] = data[outside][inside][3]
                    sample_row += 1

            else:
                for inside in range(len(data[outside])):
                    if internal_counter == 0:
                        # Target Name
                        target_name = data[outside][inside][2]
                        ws[f"{alphabet[7]}{right_current_row}"] = target_name
                        ws[f"{alphabet[7]}{right_current_row}"].font = Font(bold=True)

                        for i in range(num_labels):
                            cell_letter = i + 8
                            cell_num = 1 + right_current_row
                            ws[f"{alphabet[cell_letter]}{cell_num}"] = labels[i]
                            ws[f"{alphabet[cell_letter]}{cell_num}"].font = Font(bold=True)

                        right_current_row += len(data[outside]) + 4

                        if left_current_row < right_current_row:
                            left_current_row = right_current_row
                        elif left_current_row > right_current_row:
                            right_current_row = left_current_row

                        # +1 for spacing
                        last_hori_cell = [cell_letter + 1, cell_num]

                        sample_col = last_hori_cell[0] - num_labels
                        sample_row = last_hori_cell[1] + 1

                    # establish adjusted input cell location so that you can reference there while making the rest
                    if internal_counter == 0:
                        make_cell(sample_col, sample_row + len(data[outside]), "Adjusted Input (100%)")
                        adj_input_coor = [sample_col + 1, sample_row + len(data[outside])]

                    current_sample = data[outside][inside][1]

                    if inside + 1 < len(data[outside]):
                        next_sample = data[outside][inside + 1][1]
                    else:
                        next_sample = "!@#&I#!@^&*!@#*&^!@#()*&"

                    if next_sample != current_sample:
                        # Sample Name
                        sample_name = data[outside][inside][1]
                        ws[f"{alphabet[sample_col]}{sample_row - counter}"] = sample_name
                        # Average
                        ws[
                            f"{alphabet[sample_col + 2]}{sample_row - counter}"] = f"=AVERAGE({alphabet[sample_col + 1]}{sample_row - counter}: {alphabet[sample_col + 1]}{sample_row})"
                        # deltact
                        ws[
                            f"{alphabet[sample_col + 3]}{sample_row - counter}"] = f"={alphabet[adj_input_coor[0]]}{adj_input_coor[1]}- {alphabet[sample_col + 2]}{sample_row - counter}"

                        # % Input
                        make_cell(sample_col + 4, sample_row - counter,
                                  f"=100*2^{alphabet[sample_col + 3]}{sample_row - counter}")
                        percent_inputs.append([sample_name, target_name, [sample_col + 4, sample_row - counter]])

                        # now add values to adj input
                        for ref in reference_targets:
                            if ref == data[outside][inside][1]:
                                ave_input = [sample_col + 2, sample_row - counter]
                                make_cell(adj_input_coor[0], adj_input_coor[1],
                                          f"={alphabet[ave_input[0]]}{ave_input[1]}-{alphabet[-2]}3")
                        counter = 0
                    else:
                        counter += 1

                    internal_counter += 1

                    # Raw
                    ws[f"{alphabet[sample_col + 1]}{sample_row}"] = data[outside][inside][3]
                    sample_row += 1
                spacing += 1

        # # Make mini table to plot graphs
        # Create lists of data to plot
        graph_data = []
        temp_data = []

        for inp in percent_inputs:
            if inp[0] in target:
                temp_data.append(inp)
        graph_data.append(temp_data)

        # sort data according to target name, hopefully this will make it so that same targets are compared for the chart
        # might be a source of error and may need better sorting e.g. nested list
        graph_data = sorted(graph_data[0], key=lambda x: x[1])

        row_num_table = 3
        col_num_table = 14
        table_row_count = 0
        table_col_count = 0
        prev_target = ""
        min_col = 0
        min_row = 0
        extra_sample_label = []

        for num in range(len(graph_data)):
            # Sample label
            sample_name = graph_data[num][0]
            make_cell(col_num_table + 1, row_num_table + table_row_count, f"{graph_data[num][0]}", bold=True)
            if num == 0:
                min_col = col_num_table + 1 + 1
                min_row = row_num_table + table_row_count
            extra_sample_label = [col_num_table + 1, row_num_table + table_row_count + 1]
            current_target = graph_data[num][1]
            if current_target not in prev_target:
                table_col_count += 1
                table_row_count = 0

                # Target label
                make_cell(col_num_table + table_col_count + 1, row_num_table + table_row_count - 1, current_target,
                          bold=True)
                # % Input
                make_cell(col_num_table + table_col_count + 1, row_num_table + table_row_count,
                          f"={alphabet[graph_data[num][2][0]]}{graph_data[num][2][1]}")

                table_row_count += 1
                prev_target = current_target
            else:
                # % Input
                make_cell(col_num_table + table_col_count + 1, row_num_table + table_row_count,
                          f"={alphabet[graph_data[num][2][0]]}{graph_data[num][2][1]}")

                table_row_count += 1
                prev_target = current_target

        make_cell(extra_sample_label[0], extra_sample_label[1], "")
        max_col = col_num_table + table_col_count + 1 + 2
        max_row = row_num_table + table_row_count

        # Make graphs
        chart = BarChart()
        chart.overlap = -10
        chart.title = f"{sample_name.split()[1]} % Input Change"
        chart.y_axis.title = f"% input of {sample_name.split()[1]}"
        chart.x_axis.title = "Sample"

        values = Reference(ws, min_col=min_col + 1, min_row=min_row - 1, max_col=max_col - 1, max_row=max_row - 1)
        graph_labels = Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row - 1)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(graph_labels)

        ws.add_chart(chart, f"{alphabet[col_num_table]}{row_num_table + table_row_count}")

        table_row_count += 17

        wb.save(output_filename)
