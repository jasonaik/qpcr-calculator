import json
import os
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import NumDataSource, NumRef
from openpyxl.chart.error_bar import ErrorBars
from openpyxl.styles import Font, PatternFill, Alignment
import xlrd
from collections import OrderedDict
from openpyxl.utils import get_column_letter


'''
Rules:
1. Need to label everything correctly
2. Gene target names need to be consistent with previous inputs
3. Control sample names have to be under Control Group
'''


def default_settings():
    """Return the default settings."""
    return {
        "REF_MODE": "Exclude Other References",
        "DECIMAL_PLACES": 100,
        "BATCH_TEXT_COLOR": "ff003822",
        "BATCH_FILL_COLOR": "ffedffdc",
        "AVERAGE_BATCHES_NAME": "Average (Do not delete this if you want continuous input)",
        "PERCENTAGE_INPUT": 1,
        "SAMPLE_COLUMN": ["sample", "sample name", "name"],
        "WELL_COLUMN": ["well"],
        "GENE_COLUMN": ["gene", "target", "target name", "detector", "primer/probe", "primer", "probe"],
        "CT_COLUMN": ["ct", "cq", "cт"]
    }


def get_settings():
    """Load settings from JSON file. If the file doesn't exist, return default settings."""
    try:
        with open("settings.json", "r") as file:
            return json.load(file)
    except FileNotFoundError:
        return default_settings()


settings = get_settings()

SAMPLE_COLUMN = settings["SAMPLE_COLUMN"]
WELL_COLUMN = settings["WELL_COLUMN"]
GENE_COLUMN = settings["GENE_COLUMN"]
CT_COLUMN = settings["CT_COLUMN"]

AVERAGE_BATCHES_NAME = settings["AVERAGE_BATCHES_NAME"]
BATCH_TEXT_COLOR = settings["BATCH_TEXT_COLOR"]
BATCH_FILL_COLOR = settings["BATCH_FILL_COLOR"]
DECIMAL_PLACES = settings["DECIMAL_PLACES"]

if settings["REF_MODE"] == "Include Other References":
    REF_MODE = 1
else:
    REF_MODE = 0

# SAMPLE_COLUMN = ["sample", "sample name", "name"]
# WELL_COLUMN = ["well"]
# GENE_COLUMN = ["gene", "target", "target name", "detector", "primer/probe", "primer", "probe"]
# CT_COLUMN = ["ct", "cq", "cт"]
#
# AVERAGE_BATCHES_NAME = "Average (Do not delete this if you want continuous input)"
# BATCH_TEXT_COLOR = "ff003822"
# BATCH_FILL_COLOR = "ffedffdc"
# DECIMAL_PLACES = 100
# REF_MODE = 0

class InvalidExcelLayoutException(Exception):
    pass


class InvalidDataLayoutException(Exception):
    pass


def get_data(wb_file):
    label_row = -1
    data = []
    targets = []
    samples = []

    file_extension = os.path.splitext(wb_file)[1]

    if file_extension == '.xls':
        wb = xlrd.open_workbook(wb_file)
        read_ws = wb.sheet_by_index(0)
        row_count, col_count = read_ws.nrows, read_ws.ncols

        for row_num in range(0, row_count):
            cell = read_ws.cell(row_num, 0)
            if cell.value.lower() in WELL_COLUMN:
                label_row = row_num
                for num1 in range(0, col_count):
                    cell = read_ws.cell(label_row, num1)
                    if cell.value.lower() in SAMPLE_COLUMN:
                        for i in range(label_row + 1, row_count):
                            cell = read_ws.cell(i, num1)
                            sample_name = cell.value
                            data.append([sample_name])

                    if cell.value.lower() in GENE_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count):
                            cell = read_ws.cell(i, num1)
                            target_name = cell.value
                            data[counter].append(target_name)
                            counter += 1

                    if cell.value.lower() in CT_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count):
                            cell = read_ws.cell(i, num1)
                            ct = cell.value
                            data[counter].append(ct)
                            counter += 1
        # Replace the generic Exception with the custom one
        if label_row == -1:
            raise InvalidExcelLayoutException("Invalid Excel File Layout for file: {}".format(wb_file))

    elif file_extension == '.xlsx':
        wb = openpyxl.load_workbook(wb_file)
        read_ws = wb.active
        row_count = read_ws.max_row
        col_count = read_ws.max_column

        for row_num in range(1, row_count + 1):  # openpyxl is 1-indexed
            cell = read_ws.cell(row=row_num, column=1)
            if cell.value and cell.value.lower() in WELL_COLUMN:
                label_row = row_num
                for num1 in range(1, col_count + 1):
                    cell = read_ws.cell(row=label_row, column=num1)
                    if cell.value and cell.value.lower() in SAMPLE_COLUMN:
                        for i in range(label_row + 1, row_count + 1):
                            cell = read_ws.cell(row=i, column=num1)
                            sample_name = cell.value
                            data.append([sample_name])

                    if cell.value and cell.value.lower() in GENE_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count + 1):
                            cell = read_ws.cell(row=i, column=num1)
                            target_name = cell.value
                            data[counter].append(target_name)
                            counter += 1

                    if cell.value and cell.value.lower() in CT_COLUMN:
                        counter = 0
                        for i in range(label_row + 1, row_count + 1):
                            cell = read_ws.cell(row=i, column=num1)
                            ct = cell.value
                            data[counter].append(ct)
                            counter += 1

        # Replace the generic Exception with the custom one
        if label_row == -1:
            raise InvalidExcelLayoutException("Invalid Excel File Layout for file: {}".format(wb_file))


    else:
        raise ValueError(f"Unsupported file extension: {file_extension}")

    # clean data
    for point in data.copy():
        if point[0] == "":
            point[0] = "Unknown Sample"

        if point[2] == "" or point[2] is None:
            data.remove(point)

    junk = []
    for i in range(len(data)):
        if len(data[i]) < 3:
            junk.append(data[i])

    for i in range(len(junk)):
        data.remove(junk[i])

    for datum in data:
        if datum[0] == "":
            messagebox.showerror("Error",
                                 "One or more of the wells do not have a target, "
                                 "calculations may not be correct, "
                                 "you can amend the export spreadsheet or export "
                                 "a new spreadsheet that has the correct labels"
                                 "Could try updating spreadsheet column names in advanced settings")

    for datum in data:
        sample_name = datum[0]
        target_name = datum[1]
        if sample_name not in samples:
            samples.append(sample_name)
        if target_name not in targets:
            targets.append(target_name)

    return data, samples, targets


# multiple_ref_mode modes: 0 (separate, exclude other refs), 1 (separate, include other refs),
# modes not included yet: 2 (arithmatic mean), 3 (geometric mean)
def write_wb(data, reference_targets: list, fold_change_targets: list, output_filename, first_time=True,
             decimal_places=DECIMAL_PLACES, multiple_ref_mode=REF_MODE):
    def make_cell(col_num, row_num, input_val, bold=False):
        if isinstance(input_val, int):
            ws[f"{get_column_letter(col_num)}{row_num}"] = f"={input_val}"
        else:
            ws[f"{get_column_letter(col_num)}{row_num}"] = f"{input_val}"

        if bold:
            bold_cell = ws[f"{get_column_letter(col_num)}{row_num}"]
            bold_cell.font = Font(bold=True)

    def color_all_rows(color_code, row_num):
        fill_color = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
        for col in range(1, 100):
            ws.cell(row=row_num, column=col).fill = fill_color

    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
    labels_target = ["Target", "Raw", "Average", "delta Ct", "Fold Change"]
    cell_letter, cell_num = 0, 0

    samples = []
    targets = []

    # This is a bad idea but I'm too lazy to change it lol
    data_dict = OrderedDict()

    for datum in data:
        if datum[0] not in samples:
            samples.append(datum[0])
        if datum[1] not in targets:
            targets.append(datum[1])

    # Make undetected = 0
    for datum in data:
        if not isinstance(datum[2], float) and not isinstance(datum[2], int):
            datum[2] = ""
    # Sort data according to samples
    data = sorted(data, key=lambda x: x[1])
    # Create lists for data
    for num in range(len(samples)):
        data_dict.update({samples[num]: []})
    for num in range(len(data)):
        data_dict[data[num][0]].append(data[num])

    # Convert items in data_dict to lists
    dict_list = [[key, value] for key, value in data_dict.items()]

    # Remove unselected comboboxes
    fold_change_targets = [i for i in fold_change_targets if len(i) > 0]

    control_treatment = [[target[0], target[1]] for target in fold_change_targets]

    # Makes it so that combos with same controls are grouped together
    # Group according to the first element of the list
    grouped_dict = {}

    for item in control_treatment:
        if item[0] in grouped_dict:
            grouped_dict[item[0]].append(item[1])
        else:
            grouped_dict[item[0]] = [item[0], item[1]]

    split_controls = list(grouped_dict.values())

    org_dict_list = [[] for control in split_controls]

    for num in range(len(split_controls)):
        for num1 in range(len(split_controls[num])):
            for i in range(len(dict_list)):
                if split_controls[num][num1] == dict_list[i][0]:
                    org_dict_list[num].append(dict_list[i])

    if first_time:
        wb = Workbook()

        # Remove the default "Sheet"
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)
    else:
        wb = openpyxl.load_workbook(output_filename)

    for ref in reference_targets:
        final_data = []
        other_refs = reference_targets.copy()
        if multiple_ref_mode == 0:
            other_refs.remove(ref)
        else:
            other_refs = []

        # So that reference targets are always at the back.
        for outer_idx in range(len(org_dict_list)):
            for inner_idx in range(len(org_dict_list[outer_idx])):
                data_list = org_dict_list[outer_idx][inner_idx][1]

                # Separate lists based on ref
                non_ref_items = [item for item in data_list if item[1] != ref]
                ref_items = [item for item in data_list if item[1] == ref]

                # Combine the lists again
                org_dict_list[outer_idx][inner_idx][1] = non_ref_items + ref_items

        for num in range(len(org_dict_list)):
            final_data.append([])
            for num_sample in range(len(org_dict_list[num])):
                final_data[num].append([org_dict_list[num][num_sample][0]])
                temp = []
                for datum in org_dict_list[num][num_sample][1]:
                    if datum[1] not in other_refs:
                        temp.append(datum)
                final_data[num][num_sample].append(temp)

        # # # purge combos that won't work with current ref
        # # Case 1 - ran out of genes, Case 2 - don't have current gene
        filtered_data = []
        for datum in final_data:
            # Check if any inner datum i has fewer than 2 unique targets and contains the specific target.
            if all(len(set(x[1] for x in i[1])) >= 2 and ref in [x[1] for x in i[1]] for i in datum):
                filtered_data.append(datum)

        for exp in filtered_data:

            # Might need to change this to work with continuous data with different sample names
            if first_time:
                ws = wb.create_sheet(f"{exp[0][0]} - {ref}")  # Create a new sheet for each exp
            else:
                ws = wb[f"{exp[0][0]} - {ref}"]

            current_batch = None
            average_cell = None

            for cell in range(1, ws.max_row + 1):
                current_cell = ws[f"A{cell}"]
                if current_cell.value:
                    if current_cell.value == AVERAGE_BATCHES_NAME:
                        average_cell = current_cell
                    if "Batch" in current_cell.value:
                        current_batch = current_cell.value.split()[1]

            even_rows = 0
            odd_rows = 0
            for num in range(len(exp)):
                if num % 2 == 0:
                    even_rows += len(exp[num][1]) + 3
                else:
                    odd_rows += len(exp[num][1]) + 3
            minimum_rows = 15 + len(exp) + 2
            num_rows_inserted = max(even_rows, odd_rows, minimum_rows)

            average_samples = [sample[0] for sample in exp]

            if current_batch is None:
                current_batch = 0
                next_batch_row = 1
                color_all_rows(BATCH_FILL_COLOR, 1)
                make_cell(1, 1, "Batch 1")
                ws["A1"].font = Font(italic=True, bold=True, color=BATCH_TEXT_COLOR)

                left_current_row = 2
                right_current_row = 2

                # Color rows for separation
                average_starting_row = num_rows_inserted + 2
                color_all_rows(BATCH_FILL_COLOR, average_starting_row)
                make_cell(1, average_starting_row, AVERAGE_BATCHES_NAME)
                ws[f"A{average_starting_row}"].font = Font(italic=True, bold=True, color=BATCH_TEXT_COLOR)
                average_cell = ws[f"A{average_starting_row}"]

            else:
                next_batch_row = average_cell.row
                merged_ranges_to_unmerge = list(ws.merged_cells.ranges)
                target_row = list(ws.iter_rows(min_row=average_cell.row + 1, max_row=average_cell.row + 1))[0]
                ave_sd_row = list(ws.iter_rows(min_row=average_cell.row + 1 + (len(average_samples) + 3),
                                               max_row=average_cell.row + 1 + (len(average_samples) + 3)))[0]

                for cell in target_row:
                    for merged_range in merged_ranges_to_unmerge:
                        if cell.coordinate in merged_range:
                            if str(merged_range) in ws.merged_cells:
                                ws.unmerge_cells(str(merged_range))
                            merged_ranges_to_unmerge.remove(merged_range)

                for cell in ave_sd_row:
                    for merged_range in merged_ranges_to_unmerge:
                        if cell.coordinate in merged_range:
                            if str(merged_range) in ws.merged_cells:
                                ws.unmerge_cells(str(merged_range))
                            merged_ranges_to_unmerge.remove(merged_range)

                # Get fresh cell references after unmerging
                target_row = list(ws.iter_rows(min_row=average_cell.row + 1, max_row=average_cell.row + 1))[0]
                ave_sd_row = list(ws.iter_rows(min_row=average_cell.row + 1 + (len(average_samples) + 3),
                                               max_row=average_cell.row + 1 + (len(average_samples) + 3)))[0]

                for cell in target_row:
                    ws[f"{cell.column_letter}{cell.row}"].alignment = Alignment(horizontal='general', vertical='bottom')

                for cell in ave_sd_row:
                    ws[f"{cell.column_letter}{cell.row}"].alignment = Alignment(horizontal='general', vertical='bottom')

                ws.insert_rows(idx=average_cell.row, amount=num_rows_inserted + 1)

                color_all_rows(BATCH_FILL_COLOR, next_batch_row)
                make_cell(1, next_batch_row, f"Batch {int(current_batch) + 1}")
                ws[f"A{next_batch_row}"].font = Font(italic=True, bold=True, color=BATCH_TEXT_COLOR)

                left_current_row = next_batch_row + 1
                right_current_row = next_batch_row + 1

            spacing = 1
            deltacts = []

            for outside in range(len(exp)):
                num_targets = len(exp[outside][1])
                if outside % 2 == 0:
                    # Sample Name
                    sample_name = exp[outside][0]
                    ws[f"{alphabet[1]}{left_current_row}"] = exp[outside][0]
                    ws[f"{alphabet[1]}{left_current_row}"].font = Font(bold=True)
                    ws.merge_cells(f"{alphabet[1]}{left_current_row}:{alphabet[4]}{left_current_row}")
                    ws[f"{alphabet[1]}{left_current_row}"].alignment = Alignment(horizontal='center', vertical='center')

                    for i in range(4):
                        cell_letter = i + 1
                        cell_num = 1 + left_current_row
                        ws[f"{alphabet[cell_letter]}{cell_num}"] = labels_target[i]
                        ws[f"{alphabet[cell_letter]}{cell_num}"].font = Font(bold=True)

                    left_current_row += len(exp[outside][1]) + 3

                    # +1 for spacing
                    last_hori_cell = [cell_letter + 1, cell_num]
                    sample_col = last_hori_cell[0] - 4
                    sample_row = last_hori_cell[1] + 1

                    initial_apt_samples = [datum for datum in exp[outside]]
                    apt_samples = [initial_apt_samples[1]]

                    ctrl_first_samples_double = apt_samples.copy()
                    ctrl_first_samples = ctrl_first_samples_double[0].copy()
                    ctrl_first_samples.reverse()

                    num_refs = 0
                    ref_gene = None
                    for num in range(len(ctrl_first_samples)):
                        if num == 0:
                            ref_gene = ctrl_first_samples[num][1]
                            num_refs += 1
                        elif num > 0 and ref_gene == ctrl_first_samples[num][1]:
                            num_refs += 1

                    counter = 0
                    for num in range(len(apt_samples[0])):
                        current_target = apt_samples[0][num][1]
                        if num + 1 < len(apt_samples[0]):
                            next_target = apt_samples[0][num + 1][1]
                        else:
                            next_target = "!@#&I#!@^&*!@#*&^!@#()*&"

                        if next_target != current_target:
                            target_name = apt_samples[0][num][1]

                            ws[f"{alphabet[sample_col]}{sample_row - counter}"] = target_name
                            # Average
                            ws[
                                f"{alphabet[sample_col + 2]}{sample_row - counter}"] = f"=ROUND(AVERAGE({alphabet[sample_col + 1]}{sample_row - counter}: {alphabet[sample_col + 1]}{sample_row}), {decimal_places})"
                            # deltact
                            ws[
                                f"{alphabet[sample_col + 3]}{sample_row - counter}"] = f"=ROUND({alphabet[sample_col + 2]}{sample_row - counter}-{alphabet[sample_col + 2]}{last_hori_cell[1] + num_targets - (num_refs - 1)}, {decimal_places})"

                            deltacts.append(
                                [sample_name, target_name, f"{alphabet[sample_col + 3]}{sample_row - counter}"])
                            counter = 0
                        else:
                            counter += 1

                        # Raw
                        if isinstance(apt_samples[0][num][2], int) or isinstance(apt_samples[0][num][2], float):
                            ws[f"{alphabet[sample_col + 1]}{sample_row}"] = round(apt_samples[0][num][2],
                                                                                  decimal_places)

                        sample_row += 1

                else:
                    sample_name = exp[outside][0]
                    ws[f"{alphabet[7]}{right_current_row}"] = exp[outside][0]
                    ws[f"{alphabet[7]}{right_current_row}"].font = Font(bold=True)
                    ws.merge_cells(f"{alphabet[7]}{right_current_row}:{alphabet[10]}{right_current_row}")
                    ws[f"{alphabet[7]}{right_current_row}"].alignment = Alignment(horizontal='center',
                                                                                  vertical='center')

                    # Labels
                    for i in range(4):
                        cell_letter = i + 7
                        cell_num = 1 + right_current_row
                        ws[f"{alphabet[cell_letter]}{cell_num}"] = labels_target[i]
                        ws[f"{alphabet[cell_letter]}{cell_num}"].font = Font(bold=True)

                    right_current_row += len(exp[outside][1]) + 3

                    if left_current_row < right_current_row:
                        left_current_row = right_current_row
                    elif left_current_row > right_current_row:
                        right_current_row = left_current_row

                    # +1 for spacing
                    last_hori_cell = [cell_letter, cell_num]
                    sample_col = last_hori_cell[0] - 4 + 1
                    sample_row = last_hori_cell[1] + 1

                    initial_apt_samples = [datum for datum in exp[outside]]
                    apt_samples = [initial_apt_samples[1]]

                    ctrl_first_samples_double = apt_samples.copy()
                    ctrl_first_samples = ctrl_first_samples_double[0].copy()
                    ctrl_first_samples.reverse()

                    num_refs = 0
                    ref_gene = None
                    for num in range(len(ctrl_first_samples)):
                        if num == 0:
                            ref_gene = ctrl_first_samples[num][1]
                            num_refs += 1
                        elif num > 0 and ref_gene == ctrl_first_samples[num][1]:
                            num_refs += 1

                    counter = 0
                    for num in range(len(apt_samples[0])):
                        current_target = apt_samples[0][num][1]
                        if num + 1 < len(apt_samples[0]):
                            next_target = apt_samples[0][num + 1][1]
                        else:
                            next_target = "!@#&I#!@^&*!@#*&^!@#()*&"

                        if next_target != current_target:
                            target_name = apt_samples[0][num][1]

                            ws[f"{alphabet[sample_col]}{sample_row - counter}"] = target_name
                            # Average
                            ws[
                                f"{alphabet[sample_col + 2]}{sample_row - counter}"] = f"=ROUND(AVERAGE({alphabet[sample_col + 1]}{sample_row - counter}: {alphabet[sample_col + 1]}{sample_row}), {decimal_places})"
                            # deltact
                            ws[
                                f"{alphabet[sample_col + 3]}{sample_row - counter}"] = f"=ROUND({alphabet[sample_col + 2]}{sample_row - counter}-{alphabet[sample_col + 2]}{last_hori_cell[1] + num_targets - (num_refs - 1)}, {decimal_places})"

                            deltacts.append(
                                [sample_name, target_name, f"{alphabet[sample_col + 3]}{sample_row - counter}"])
                            counter = 0
                        else:
                            counter += 1

                        # Raw
                        if isinstance(apt_samples[0][num][2], int) or isinstance(apt_samples[0][num][2], float):
                            ws[f"{alphabet[sample_col + 1]}{sample_row}"] = round(apt_samples[0][num][2],
                                                                                  decimal_places)

                        sample_row += 1

                    spacing += 1

            # Current batch row
            next_batch_row += 1

            # Fold change label
            make_cell(15, next_batch_row, "Expression Fold Change", bold=True)
            ws.merge_cells(f"O{next_batch_row}:S{next_batch_row}")
            centered_alignment = Alignment(horizontal='center', vertical='center')
            ws[f'O{next_batch_row}'].alignment = centered_alignment

            # # Generate nested list for fold change targets
            org_coor = [[target[0], [], target[1], []] for target in fold_change_targets if target[0] == exp[0][0]]
            for ct in deltacts:
                for num in range(len(org_coor)):
                    if ct[0] == org_coor[num][0] and ct[0]:
                        org_coor[num][1].append(ct)
                    elif ct[0] == org_coor[num][2]:
                        org_coor[num][3].append(ct)

            row_num_table = next_batch_row + 2
            col_num_table = 15
            table_row_count = 0
            min_col = 0
            max_col = 0
            min_row = 0
            max_row = 0

            # Plot chart

            for num in range(len(org_coor)):
                if num == 0:
                    min_col = col_num_table
                    min_row = row_num_table + table_row_count
                for num1 in range(len(org_coor[num][1])):
                    control_list = org_coor[num][1][num1]
                    treatment_list = org_coor[num][3][num1]
                    if num1 == 0:
                        # treatment sample name
                        make_cell(col_num_table, row_num_table + table_row_count + 1, treatment_list[0], bold=True)

                    if num == 0:
                        # Control Sample name
                        make_cell(col_num_table, row_num_table + table_row_count, control_list[0], bold=True)
                        # Control Fold Change
                        make_cell(col_num_table + num1 + 1, row_num_table + table_row_count, 1)
                        # target names
                        if ws[
                            f"{alphabet[col_num_table + num1 + 1]}{row_num_table + table_row_count - 1}"].value is None:
                            make_cell(col_num_table + num1 + 1, row_num_table + table_row_count - 1,
                                      control_list[1], bold=True)

                    # Treatment Fold Change
                    make_cell(col_num_table + num1 + 1, row_num_table + table_row_count + 1,
                              f"=ROUND(2^-({treatment_list[2]}-{control_list[2]}), {decimal_places})")

                    max_col = col_num_table + num1 + 1
                    max_row = row_num_table + table_row_count + 2
                table_row_count += 1

            # Make graph for current data
            chart = BarChart()
            chart.overlap = -10
            chart.title = f"Expression Fold Change"
            chart.y_axis.title = "Expression Fold Change"
            chart.x_axis.title = "Sample"

            values = Reference(ws, min_col=min_col + 1, min_row=min_row - 1, max_col=max_col - 1, max_row=max_row - 1)
            labels = Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row - 1)
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(labels)

            ws.add_chart(chart, f"{alphabet[col_num_table - 1]}{row_num_table + table_row_count + 1}")

            table_row_count += 17

            def copy_and_paste(min_row_cp, max_row_cp, min_col_cp, max_col_cp, target_start_col, target_start_row):

                # Define the source range based on min/max rows and columns
                source_range = ws.iter_rows(min_row=min_row_cp, max_row=max_row_cp, min_col=min_col_cp,
                                            max_col=max_col_cp)
                row_counter = 0
                for row in source_range:
                    col_counter = 0
                    row_counter += 1
                    for cell in row:
                        try:
                            col_counter += 1
                            # Copy values
                            target_cell = ws.cell(row=target_start_row + row_counter,
                                                  column=target_start_col + col_counter)
                            target_cell.value = cell.value

                            # Copy bold formatting
                            if cell.font and cell.font.bold:
                                target_cell.font = openpyxl.styles.Font(bold=True)

                            if target_cell.value == "Expression Fold Change":
                                target_cell.value = f"Batch {int(current_batch) + 1}"

                        # AttributeError: 'MergedCell' object attribute 'value' is read-only
                        except AttributeError:
                            pass

            # Batch Expression Fold Change
            # Honestly could have just repeated what you did for the expression fold change tables LOL
            col_location = None
            none_counter = 0
            for cell in list(ws.iter_rows(min_row=average_cell.row + 3, max_row=average_cell.row + 3))[0]:
                if cell.value is None:
                    none_counter += 1
                    if none_counter == 2:
                        col_location = cell.column
                        break
                else:
                    none_counter = 0

            copy_and_paste(min_row_cp=min_row - 2, max_row_cp=max_row, min_col_cp=min_col, max_col_cp=max_col - 1,
                           target_start_col=col_location - 1, target_start_row=average_cell.row)

            # # For Average, moved average_samples up to work with unmerging
            average_targets = []
            for target in exp[0][1]:
                if target[1] not in average_targets:
                    average_targets.append(target[1])
            average_targets.remove(ref)

            average_table_col = 2
            sd_table_col = 7
            average_table_row = average_cell.row + len(average_samples) + 4

            make_cell(col_num=average_table_col, row_num=average_table_row, input_val="Average", bold=True)
            make_cell(col_num=sd_table_col, row_num=average_table_row, input_val="Standard Deviation", bold=True)

            # Target Labels
            for num in range(len(average_targets)):
                make_cell(average_table_col + num + 1, average_table_row + 1, average_targets[num], bold=True)
                make_cell(sd_table_col + num + 1, average_table_row + 1, average_targets[num], bold=True)

            # Samples Labels
            for num in range(len(average_samples)):
                make_cell(average_table_col, average_table_row + num + 2, average_samples[num], bold=True)
                make_cell(sd_table_col, average_table_row + num + 2, average_samples[num], bold=True)

            next_batch_num = int(current_batch) + 1

            # Fill in values of average and sd
            ave_min_col = average_table_col + 1
            ave_max_col = 0

            ave_sd_min_row = average_table_row + 2
            ave_sd_max_row = 0

            for num_target in range(len(average_targets)):
                for num_sample in range(len(average_samples)):
                    ave_str = ""
                    for num in range(next_batch_num):
                        current_col = average_table_col + num_target + 1 + num * (len(average_targets) + 2)
                        current_row = average_table_row + num_sample + 2 - (len(average_samples) + 3)
                        if num == 0:
                            ave_str += f"{get_column_letter(current_col)}{current_row}"
                        else:
                            ave_str += f",{get_column_letter(current_col)}{current_row}"

                    make_cell(average_table_col + num_target + 1, average_table_row + num_sample + 2,
                              f"=AVERAGE({ave_str})")
                    make_cell(sd_table_col + num_target + 1, average_table_row + num_sample + 2, f"=STDEV({ave_str})")
                    if num_sample == (len(average_samples) - 1):
                        ave_sd_max_row = average_table_row + num_sample + 2
                if num_target == (len(average_targets) - 1):
                    ave_max_col = average_table_col + num_target + 1

            # To merge and center all Batch x, Average, Standard Deviation cells
            none_counter = 0
            value_counter = 0
            ave_sd_counter = 0
            first_col = None
            batch_num_row = list(ws.iter_rows(min_row=average_cell.row + 3, max_row=average_cell.row + 3))[0]
            for num in range(len(batch_num_row)):
                if num == 0:
                    pass
                elif batch_num_row[num].value is None:
                    value_counter = 0
                    none_counter += 1
                    if none_counter == 2:
                        break
                    ws.merge_cells(
                        f"{first_col}{batch_num_row[num].row - 2}:{batch_num_row[num - 1].column_letter}{batch_num_row[num].row - 2}")
                    ws[f'{first_col}{batch_num_row[num].row - 2}'].alignment = Alignment(horizontal='center',
                                                                                         vertical='center')
                    if ave_sd_counter < 2:
                        ave_sd_row = batch_num_row[num].row - 2 + (len(average_samples) + 3)
                        ws.merge_cells(f"{first_col}{ave_sd_row}:{batch_num_row[num - 1].column_letter}{ave_sd_row}")
                        ws[f'{first_col}{ave_sd_row}'].alignment = Alignment(horizontal='center', vertical='center')
                        ave_sd_counter += 1

                elif batch_num_row[num].value is not None and value_counter == 0:
                    value_counter += 1
                    first_col = batch_num_row[num].column_letter
                    none_counter = 0
                else:
                    value_counter += 1
                    none_counter = 0

            if current_batch == 0:
                # Make graph for average
                chart = BarChart()
                chart.overlap = -10
                chart.title = f"Average of Batches"
                chart.y_axis.title = "Expression Fold Change"
                chart.x_axis.title = "Sample"

                values = Reference(ws, min_col=ave_min_col, min_row=ave_sd_min_row - 1, max_col=ave_max_col,
                                   max_row=ave_sd_max_row)
                labels = Reference(ws, min_col=ave_min_col - 1, min_row=ave_sd_min_row, max_row=ave_sd_max_row + 1)
                chart.add_data(values, titles_from_data=True)
                chart.set_categories(labels)

                ws.add_chart(chart, f"{get_column_letter(ave_min_col - 1)}{ave_sd_max_row + 1}")
            else:

                average_chart = ws._charts[1]

                # Remove the previous data series from the chart
                average_chart.series = []

                # Add the updated data series
                values = Reference(ws, min_col=ave_min_col, min_row=ave_sd_min_row - 1, max_col=ave_max_col,
                                   max_row=ave_sd_max_row)
                labels = Reference(ws, min_col=ave_min_col - 1, min_row=ave_sd_min_row, max_row=ave_sd_max_row + 1)
                average_chart.add_data(values, titles_from_data=True)
                data_series = average_chart.series
                average_chart.set_categories(labels)

                # Reference for standard deviations
                sd_values = Reference(ws, min_col=ave_max_col + 3, min_row=ave_sd_min_row, max_col=ave_max_col * 2,
                                      max_row=ave_sd_max_row)

                # Add error bars
                num_data_cols = ave_max_col - ave_min_col + 1
                for i in range(num_data_cols):
                    # Extract the corresponding standard deviation values for the current data column from the already defined sd_values
                    sd_ref = Reference(ws, min_col=sd_values.min_col + i, min_row=sd_values.min_row,
                                       max_col=sd_values.min_col + i, max_row=sd_values.max_row)
                    ebars_numdatasource = NumDataSource(NumRef(sd_ref))
                    data_series[i].errBars = ErrorBars(errDir='y', errValType='cust', plus=ebars_numdatasource,
                                                       minus=ebars_numdatasource)

                # Update the anchor position
                average_chart.anchor = f"{get_column_letter(ave_min_col - 1)}{ave_sd_max_row + 1}"

            wb.save(f"{output_filename}")


# multiple_ref_mode modes: 0 (separate, exclude other refs), 1 (separate, include other refs),
# modes not included yet: 2 (arithmatic mean), 3 (geometric mean)
def write_wb_cont(data, fold_change_targets: list, output_filename, decimal_places=100):
    def make_cell(col_num, row_num, input_val, bold=False):
        if isinstance(input_val, int):
            ws[f"{get_column_letter(col_num)}{row_num}"] = f"={input_val}"
        else:
            ws[f"{get_column_letter(col_num)}{row_num}"] = f"{input_val}"

        if bold:
            bold_cell = ws[f"{get_column_letter(col_num)}{row_num}"]
            bold_cell.font = Font(bold=True)

    def color_all_rows(color_code, row_num):
        fill_color = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
        for col in range(1, 100):
            ws.cell(row=row_num, column=col).fill = fill_color

    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
    labels_target = ["Target", "Raw", "Average", "delta Ct", "Fold Change"]
    cell_letter, cell_num = 0, 0

    samples = []
    targets = []

    # This is a bad idea, but I'm too lazy to change it lol
    data_dict = OrderedDict()

    for datum in data:
        if datum[0] not in samples:
            samples.append(datum[0])
        if datum[1] not in targets:
            targets.append(datum[1])

    # Make undetected = 0
    for datum in data:
        if not isinstance(datum[2], float) and not isinstance(datum[2], int):
            datum[2] = ""
    # Sort data according to samples
    data = sorted(data, key=lambda x: x[1])
    # Create lists for data
    for num in range(len(samples)):
        data_dict.update({samples[num]: []})
    for num in range(len(data)):
        data_dict[data[num][0]].append(data[num])

    # Convert items in data_dict to lists
    dict_list = [[key, value] for key, value in data_dict.items()]

    # Remove unselected comboboxes
    filtered_fc = [sublist for sublist in fold_change_targets if not any('' in inner_list for inner_list in sublist[1])]

    # these lists only contains datapoints or labels, from here they will be linked through index
    fold_change_targets_only = [inner_list for sublist in filtered_fc for inner_list in sublist[1]]
    fold_change_sheets_only = [sublist[0] for sublist in filtered_fc]
    ref_genes = [sublist[2] for sublist in filtered_fc]

    # # idk what this does, but it seems to be redundant
    # control_treatment = [[target[0], target[1]] for target in fold_change_targets_only]

    # Makes it so that combos with same controls are grouped together
    # Group according to the first element of the list
    grouped_dict = {}

    for item in fold_change_targets_only:
        if item[0] in grouped_dict:
            grouped_dict[item[0]].append(item[1])
        else:
            grouped_dict[item[0]] = [item[0], item[1]]

    split_controls = list(grouped_dict.values())

    org_dict_list = [[] for control in split_controls]

    for num in range(len(split_controls)):
        for num1 in range(len(split_controls[num])):
            for i in range(len(dict_list)):
                if split_controls[num][num1] == dict_list[i][0]:
                    org_dict_list[num].append(dict_list[i])

    # So that reference targets are always at the back.
    for outer_idx in range(len(org_dict_list)):
        for inner_idx in range(len(org_dict_list[outer_idx])):
            data_list = org_dict_list[outer_idx][inner_idx][1]

            # Separate lists based on ref
            non_ref_items = [item for item in data_list if item[1] != ref_genes[outer_idx]]
            ref_items = [item for item in data_list if item[1] == ref_genes[outer_idx]]

            # Combine the lists again
            org_dict_list[outer_idx][inner_idx][1] = non_ref_items + ref_items

    wb = openpyxl.load_workbook(output_filename)

    # Starting writing sheets
    for sheet_num in range(len(fold_change_sheets_only)):

        ws = wb[fold_change_sheets_only[sheet_num]]

        exp = org_dict_list[sheet_num]

        current_batch = None
        average_cell = None

        for cell in range(1, ws.max_row + 1):
            current_cell = ws[f"A{cell}"]
            if current_cell.value:
                if current_cell.value == AVERAGE_BATCHES_NAME:
                    average_cell = current_cell
                if "Batch" in current_cell.value:
                    current_batch = current_cell.value.split()[1]

        if average_cell is None:
            raise InvalidDataLayoutException("Invalid data layout for continuous input")

        even_rows = 0
        odd_rows = 0
        for num in range(len(exp)):
            if num % 2 == 0:
                even_rows += len(exp[num][1]) + 3
            else:
                odd_rows += len(exp[num][1]) + 3
        minimum_rows = 15 + len(exp) + 2
        num_rows_inserted = max(even_rows, odd_rows, minimum_rows)

        average_samples = [sample[0] for sample in exp]

        if current_batch is None:
            current_batch = 0
            next_batch_row = 1
            color_all_rows(BATCH_FILL_COLOR, 1)
            make_cell(1, 1, "Batch 1")
            ws["A1"].font = Font(italic=True, bold=True, color=BATCH_TEXT_COLOR)

            left_current_row = 2
            right_current_row = 2

            # Color rows for separation
            average_starting_row = num_rows_inserted + 2
            color_all_rows(BATCH_FILL_COLOR, average_starting_row)
            make_cell(1, average_starting_row, AVERAGE_BATCHES_NAME)
            ws[f"A{average_starting_row}"].font = Font(italic=True, bold=True, color=BATCH_TEXT_COLOR)
            average_cell = ws[f"A{average_starting_row}"]

        else:
            next_batch_row = average_cell.row
            merged_ranges_to_unmerge = list(ws.merged_cells.ranges)
            target_row = list(ws.iter_rows(min_row=average_cell.row + 1, max_row=average_cell.row + 1))[0]
            ave_sd_row = list(ws.iter_rows(min_row=average_cell.row + 1 + (len(average_samples) + 3),
                                           max_row=average_cell.row + 1 + (len(average_samples) + 3)))[0]

            for cell in target_row:
                for merged_range in merged_ranges_to_unmerge:
                    if cell.coordinate in merged_range:
                        if str(merged_range) in ws.merged_cells:
                            ws.unmerge_cells(str(merged_range))
                        merged_ranges_to_unmerge.remove(merged_range)

            for cell in ave_sd_row:
                for merged_range in merged_ranges_to_unmerge:
                    if cell.coordinate in merged_range:
                        if str(merged_range) in ws.merged_cells:
                            ws.unmerge_cells(str(merged_range))
                        merged_ranges_to_unmerge.remove(merged_range)

            # Get fresh cell references after unmerging
            target_row = list(ws.iter_rows(min_row=average_cell.row + 1, max_row=average_cell.row + 1))[0]
            ave_sd_row = list(ws.iter_rows(min_row=average_cell.row + 1 + (len(average_samples) + 3),
                                           max_row=average_cell.row + 1 + (len(average_samples) + 3)))[0]

            for cell in target_row:
                ws[f"{cell.column_letter}{cell.row}"].alignment = Alignment(horizontal='general', vertical='bottom')

            for cell in ave_sd_row:
                ws[f"{cell.column_letter}{cell.row}"].alignment = Alignment(horizontal='general', vertical='bottom')

            ws.insert_rows(idx=average_cell.row, amount=num_rows_inserted + 1)

            color_all_rows(BATCH_FILL_COLOR, next_batch_row)
            make_cell(1, next_batch_row, f"Batch {int(current_batch) + 1}")
            ws[f"A{next_batch_row}"].font = Font(italic=True, bold=True, color=BATCH_TEXT_COLOR)

            left_current_row = next_batch_row + 1
            right_current_row = next_batch_row + 1

        spacing = 1
        deltacts = []

        for outside in range(len(exp)):
            num_targets = len(exp[outside][1])
            if outside % 2 == 0:
                # Sample Name
                sample_name = exp[outside][0]
                ws[f"{alphabet[1]}{left_current_row}"] = exp[outside][0]
                ws[f"{alphabet[1]}{left_current_row}"].font = Font(bold=True)
                ws.merge_cells(f"{alphabet[1]}{left_current_row}:{alphabet[4]}{left_current_row}")
                ws[f"{alphabet[1]}{left_current_row}"].alignment = Alignment(horizontal='center', vertical='center')

                for i in range(4):
                    cell_letter = i + 1
                    cell_num = 1 + left_current_row
                    ws[f"{alphabet[cell_letter]}{cell_num}"] = labels_target[i]
                    ws[f"{alphabet[cell_letter]}{cell_num}"].font = Font(bold=True)

                left_current_row += len(exp[outside][1]) + 3

                # +1 for spacing
                last_hori_cell = [cell_letter + 1, cell_num]
                sample_col = last_hori_cell[0] - 4
                sample_row = last_hori_cell[1] + 1

                initial_apt_samples = [datum for datum in exp[outside]]
                apt_samples = [initial_apt_samples[1]]

                ctrl_first_samples_double = apt_samples.copy()
                ctrl_first_samples = ctrl_first_samples_double[0].copy()
                ctrl_first_samples.reverse()

                num_refs = 0
                ref_gene = None
                for num in range(len(ctrl_first_samples)):
                    if num == 0:
                        ref_gene = ctrl_first_samples[num][1]
                        num_refs += 1
                    elif num > 0 and ref_gene == ctrl_first_samples[num][1]:
                        num_refs += 1

                counter = 0
                for num in range(len(apt_samples[0])):
                    current_target = apt_samples[0][num][1]
                    if num + 1 < len(apt_samples[0]):
                        next_target = apt_samples[0][num + 1][1]
                    else:
                        next_target = "!@#&I#!@^&*!@#*&^!@#()*&"

                    if next_target != current_target:
                        target_name = apt_samples[0][num][1]

                        ws[f"{alphabet[sample_col]}{sample_row - counter}"] = target_name
                        # Average
                        ws[
                            f"{alphabet[sample_col + 2]}{sample_row - counter}"] = f"=ROUND(AVERAGE({alphabet[sample_col + 1]}{sample_row - counter}: {alphabet[sample_col + 1]}{sample_row}), {decimal_places})"
                        # deltact
                        ws[
                            f"{alphabet[sample_col + 3]}{sample_row - counter}"] = f"=ROUND({alphabet[sample_col + 2]}{sample_row - counter}-{alphabet[sample_col + 2]}{last_hori_cell[1] + num_targets - (num_refs - 1)}, {decimal_places})"

                        deltacts.append(
                            [sample_name, target_name, f"{alphabet[sample_col + 3]}{sample_row - counter}"])
                        counter = 0
                    else:
                        counter += 1

                    # Raw
                    if isinstance(apt_samples[0][num][2], int) or isinstance(apt_samples[0][num][2], float):
                        ws[f"{alphabet[sample_col + 1]}{sample_row}"] = round(apt_samples[0][num][2],
                                                                              decimal_places)

                    sample_row += 1

            else:
                sample_name = exp[outside][0]
                ws[f"{alphabet[7]}{right_current_row}"] = exp[outside][0]
                ws[f"{alphabet[7]}{right_current_row}"].font = Font(bold=True)
                ws.merge_cells(f"{alphabet[7]}{right_current_row}:{alphabet[10]}{right_current_row}")
                ws[f"{alphabet[7]}{right_current_row}"].alignment = Alignment(horizontal='center',
                                                                              vertical='center')

                # Labels
                for i in range(4):
                    cell_letter = i + 7
                    cell_num = 1 + right_current_row
                    ws[f"{alphabet[cell_letter]}{cell_num}"] = labels_target[i]
                    ws[f"{alphabet[cell_letter]}{cell_num}"].font = Font(bold=True)

                right_current_row += len(exp[outside][1]) + 3

                if left_current_row < right_current_row:
                    left_current_row = right_current_row
                elif left_current_row > right_current_row:
                    right_current_row = left_current_row

                # +1 for spacing
                last_hori_cell = [cell_letter, cell_num]
                sample_col = last_hori_cell[0] - 4 + 1
                sample_row = last_hori_cell[1] + 1

                initial_apt_samples = [datum for datum in exp[outside]]
                apt_samples = [initial_apt_samples[1]]

                ctrl_first_samples_double = apt_samples.copy()
                ctrl_first_samples = ctrl_first_samples_double[0].copy()
                ctrl_first_samples.reverse()

                num_refs = 0
                ref_gene = None
                for num in range(len(ctrl_first_samples)):
                    if num == 0:
                        ref_gene = ctrl_first_samples[num][1]
                        num_refs += 1
                    elif num > 0 and ref_gene == ctrl_first_samples[num][1]:
                        num_refs += 1

                counter = 0
                for num in range(len(apt_samples[0])):
                    current_target = apt_samples[0][num][1]
                    if num + 1 < len(apt_samples[0]):
                        next_target = apt_samples[0][num + 1][1]
                    else:
                        next_target = "!@#&I#!@^&*!@#*&^!@#()*&"

                    if next_target != current_target:
                        target_name = apt_samples[0][num][1]

                        ws[f"{alphabet[sample_col]}{sample_row - counter}"] = target_name
                        # Average
                        ws[
                            f"{alphabet[sample_col + 2]}{sample_row - counter}"] = f"=ROUND(AVERAGE({alphabet[sample_col + 1]}{sample_row - counter}: {alphabet[sample_col + 1]}{sample_row}), {decimal_places})"
                        # deltact
                        ws[
                            f"{alphabet[sample_col + 3]}{sample_row - counter}"] = f"=ROUND({alphabet[sample_col + 2]}{sample_row - counter}-{alphabet[sample_col + 2]}{last_hori_cell[1] + num_targets - (num_refs - 1)}, {decimal_places})"

                        deltacts.append(
                            [sample_name, target_name, f"{alphabet[sample_col + 3]}{sample_row - counter}"])
                        counter = 0
                    else:
                        counter += 1

                    # Raw
                    if isinstance(apt_samples[0][num][2], int) or isinstance(apt_samples[0][num][2], float):
                        ws[f"{alphabet[sample_col + 1]}{sample_row}"] = round(apt_samples[0][num][2],
                                                                              decimal_places)

                    sample_row += 1

                spacing += 1

        # Current batch row
        next_batch_row += 1

        # Fold change label
        make_cell(15, next_batch_row, "Expression Fold Change", bold=True)
        ws.merge_cells(f"O{next_batch_row}:S{next_batch_row}")
        centered_alignment = Alignment(horizontal='center', vertical='center')
        ws[f'O{next_batch_row}'].alignment = centered_alignment

        # # Generate nested list for fold change targets

        org_coor = [[target[0], [], target[1], []] for target in fold_change_targets_only if target[0] == exp[0][0]]
        for ct in deltacts:
            for num in range(len(org_coor)):
                if ct[0] == org_coor[num][0] and ct[0]:
                    org_coor[num][1].append(ct)
                elif ct[0] == org_coor[num][2]:
                    org_coor[num][3].append(ct)

        row_num_table = next_batch_row + 2
        col_num_table = 15
        table_row_count = 0
        min_col = 0
        max_col = 0
        min_row = 0
        max_row = 0

        # Plot chart

        for num in range(len(org_coor)):
            if num == 0:
                min_col = col_num_table
                min_row = row_num_table + table_row_count
            for num1 in range(len(org_coor[num][1])):
                control_list = org_coor[num][1][num1]
                treatment_list = org_coor[num][3][num1]
                if num1 == 0:
                    # treatment sample name
                    make_cell(col_num_table, row_num_table + table_row_count + 1, treatment_list[0], bold=True)

                if num == 0:
                    # Control Sample name
                    make_cell(col_num_table, row_num_table + table_row_count, control_list[0], bold=True)
                    # Control Fold Change
                    make_cell(col_num_table + num1 + 1, row_num_table + table_row_count, 1)
                    # target names
                    if ws[
                        f"{alphabet[col_num_table + num1 + 1]}{row_num_table + table_row_count - 1}"].value is None:
                        make_cell(col_num_table + num1 + 1, row_num_table + table_row_count - 1,
                                  control_list[1], bold=True)

                # Treatment Fold Change
                make_cell(col_num_table + num1 + 1, row_num_table + table_row_count + 1,
                          f"=ROUND(2^-({treatment_list[2]}-{control_list[2]}), {decimal_places})")

                max_col = col_num_table + num1 + 1
                max_row = row_num_table + table_row_count + 2
            table_row_count += 1

        # Make graph for current data
        chart = BarChart()
        chart.overlap = -10
        chart.title = f"Expression Fold Change"
        chart.y_axis.title = "Expression Fold Change"
        chart.x_axis.title = "Sample"

        values = Reference(ws, min_col=min_col + 1, min_row=min_row - 1, max_col=max_col - 1, max_row=max_row - 1)
        labels = Reference(ws, min_col=min_col, min_row=min_row, max_row=max_row - 1)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(labels)

        ws.add_chart(chart, f"{alphabet[col_num_table - 1]}{row_num_table + table_row_count + 1}")

        table_row_count += 17

        def copy_and_paste(min_row_cp, max_row_cp, min_col_cp, max_col_cp, target_start_col, target_start_row):

            # Define the source range based on min/max rows and columns
            source_range = ws.iter_rows(min_row=min_row_cp, max_row=max_row_cp, min_col=min_col_cp,
                                        max_col=max_col_cp)
            row_counter = 0
            for row in source_range:
                col_counter = 0
                row_counter += 1
                for cell in row:
                    try:
                        col_counter += 1
                        # Copy values
                        target_cell = ws.cell(row=target_start_row + row_counter,
                                              column=target_start_col + col_counter)
                        target_cell.value = cell.value

                        # Copy bold formatting
                        if cell.font and cell.font.bold:
                            target_cell.font = openpyxl.styles.Font(bold=True)

                        if target_cell.value == "Expression Fold Change":
                            target_cell.value = f"Batch {int(current_batch) + 1}"

                    # AttributeError: 'MergedCell' object attribute 'value' is read-only
                    except AttributeError:
                        pass

        # Batch Expression Fold Change
        # Honestly could have just repeated what you did for the expression fold change tables LOL
        col_location = None
        none_counter = 0
        for cell in list(ws.iter_rows(min_row=average_cell.row + 3, max_row=average_cell.row + 3))[0]:
            if cell.value is None:
                none_counter += 1
                if none_counter == 2:
                    col_location = cell.column
                    break
            else:
                none_counter = 0

        copy_and_paste(min_row_cp=min_row - 2, max_row_cp=max_row, min_col_cp=min_col, max_col_cp=max_col - 1,
                       target_start_col=col_location - 1, target_start_row=average_cell.row)

        # # For Average, moved average_samples up to work with unmerging
        average_targets = []
        for target in exp[0][1]:
            if target[1] not in average_targets:
                average_targets.append(target[1])
        average_targets.remove(ref_genes[sheet_num])

        average_table_col = 2
        sd_table_col = 7
        average_table_row = average_cell.row + len(average_samples) + 4

        make_cell(col_num=average_table_col, row_num=average_table_row, input_val="Average", bold=True)
        make_cell(col_num=sd_table_col, row_num=average_table_row, input_val="Standard Deviation", bold=True)

        # Target Labels
        for num in range(len(average_targets)):
            make_cell(average_table_col + num + 1, average_table_row + 1, average_targets[num], bold=True)
            make_cell(sd_table_col + num + 1, average_table_row + 1, average_targets[num], bold=True)

        # Samples Labels
        for num in range(len(average_samples)):
            make_cell(average_table_col, average_table_row + num + 2, average_samples[num], bold=True)
            make_cell(sd_table_col, average_table_row + num + 2, average_samples[num], bold=True)

        next_batch_num = int(current_batch) + 1

        # Fill in values of average and sd
        ave_min_col = average_table_col + 1
        ave_max_col = 0

        ave_sd_min_row = average_table_row + 2
        ave_sd_max_row = 0

        for num_target in range(len(average_targets)):
            for num_sample in range(len(average_samples)):
                ave_str = ""
                for num in range(next_batch_num):
                    current_col = average_table_col + num_target + 1 + num * (len(average_targets) + 2)
                    current_row = average_table_row + num_sample + 2 - (len(average_samples) + 3)
                    if num == 0:
                        ave_str += f"{get_column_letter(current_col)}{current_row}"
                    else:
                        ave_str += f",{get_column_letter(current_col)}{current_row}"

                make_cell(average_table_col + num_target + 1, average_table_row + num_sample + 2,
                          f"=AVERAGE({ave_str})")
                make_cell(sd_table_col + num_target + 1, average_table_row + num_sample + 2, f"=STDEV({ave_str})")
                if num_sample == (len(average_samples) - 1):
                    ave_sd_max_row = average_table_row + num_sample + 2
            if num_target == (len(average_targets) - 1):
                ave_max_col = average_table_col + num_target + 1
                sd_max_col = sd_table_col + num_target + 1

        # To merge and center all Batch x, Average, Standard Deviation cells
        none_counter = 0
        value_counter = 0
        ave_sd_counter = 0
        first_col = None
        batch_num_row = list(ws.iter_rows(min_row=average_cell.row + 3, max_row=average_cell.row + 3))[0]
        for num in range(len(batch_num_row)):
            if num == 0:
                pass
            elif batch_num_row[num].value is None:
                value_counter = 0
                none_counter += 1
                if none_counter == 2:
                    break
                ws.merge_cells(
                    f"{first_col}{batch_num_row[num].row - 2}:{batch_num_row[num - 1].column_letter}{batch_num_row[num].row - 2}")
                ws[f'{first_col}{batch_num_row[num].row - 2}'].alignment = Alignment(horizontal='center',
                                                                                     vertical='center')
                if ave_sd_counter < 2:
                    ave_sd_row = batch_num_row[num].row - 2 + (len(average_samples) + 3)
                    ws.merge_cells(f"{first_col}{ave_sd_row}:{batch_num_row[num - 1].column_letter}{ave_sd_row}")
                    ws[f'{first_col}{ave_sd_row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ave_sd_counter += 1

            elif batch_num_row[num].value is not None and value_counter == 0:
                value_counter += 1
                first_col = batch_num_row[num].column_letter
                none_counter = 0
            else:
                value_counter += 1
                none_counter = 0

        if current_batch == 0:
            # Make graph for average
            chart = BarChart()
            chart.overlap = -10
            chart.title = f"Average of Batches"
            chart.y_axis.title = "Expression Fold Change"
            chart.x_axis.title = "Sample"

            values = Reference(ws, min_col=ave_min_col, min_row=ave_sd_min_row - 1, max_col=ave_max_col,
                               max_row=ave_sd_max_row)
            labels = Reference(ws, min_col=ave_min_col - 1, min_row=ave_sd_min_row, max_row=ave_sd_max_row + 1)
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(labels)

            ws.add_chart(chart, f"{get_column_letter(ave_min_col - 1)}{ave_sd_max_row + 1}")
        else:

            average_chart = ws._charts[1]

            # Remove the previous data series from the chart
            average_chart.series = []

            # Add the updated data series
            values = Reference(ws, min_col=ave_min_col, min_row=ave_sd_min_row - 1, max_col=ave_max_col,
                               max_row=ave_sd_max_row)
            labels = Reference(ws, min_col=ave_min_col - 1, min_row=ave_sd_min_row, max_row=ave_sd_max_row + 1)
            average_chart.add_data(values, titles_from_data=True)
            data_series = average_chart.series
            average_chart.set_categories(labels)

            # Reference for standard deviations
            sd_values = Reference(ws, min_col=ave_max_col + 3, min_row=ave_sd_min_row, max_col=ave_max_col * 2,
                                  max_row=ave_sd_max_row)

            # Add error bars
            num_data_cols = ave_max_col - ave_min_col + 1
            for i in range(num_data_cols):
                # Extract the corresponding standard deviation values for the current data column from the already defined sd_values
                sd_ref = Reference(ws, min_col=sd_values.min_col + i, min_row=sd_values.min_row,
                                   max_col=sd_values.min_col + i, max_row=sd_values.max_row)
                ebars_numdatasource = NumDataSource(NumRef(sd_ref))
                data_series[i].errBars = ErrorBars(errDir='y', errValType='cust', plus=ebars_numdatasource,
                                                   minus=ebars_numdatasource)

            # Update the anchor position
            average_chart.anchor = f"{get_column_letter(ave_min_col - 1)}{ave_sd_max_row + 1}"

        wb.save(f"{output_filename}")


# gets sample order and sheet names
def get_existing_info(file):
    wb = openpyxl.load_workbook(file)
    sheet_names = wb.sheetnames

    samples = []
    for num in range(len(sheet_names)):
        samples.append([sheet_names[num], []])
        ws = wb[sheet_names[num]]

        hit = False
        for row in range(ws.max_row, 0, -1):
            cell = ws[f"B{row}"]
            if cell.value is not None:
                hit = True
                samples[num][1].append(cell.value)
            elif cell.value is None and hit:
                break

        samples[num][1].reverse()

    return samples
